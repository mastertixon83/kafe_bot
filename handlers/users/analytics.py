import json
import os
import re

from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill

from io import BytesIO

from aiogram.dispatcher import FSMContext

from loader import dp, bot, db, logger
from aiogram import types

from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputMediaPhoto, ReplyKeyboardRemove

from states.analytics import Analytics
from utils.db_api.db_commands import DBCommands

from aiogram.dispatcher.filters import Text

db = DBCommands()

type_mailing_dict = {
    "standard": "Обычная рассылка",
    "birthday": "Предложение для именинников",
    "hall_reservation": "Призыв к бронированию",
    "shipping": "Закажите доставку",
    "loyal_card": "Владельцам карт лояльности"
}


def plural_form(n, word):
    """Определение правильного написания формы слова"""
    if 2 <= n % 10 <= 4 and (n % 100 < 10 or n % 100 >= 20):
        return f"{n} {word}а"
    else:
        return f"{n} {word}"


def data_preparation():
    """Подготовка дат для запроса"""
    # За сегодня
    start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = start_date + timedelta(days=1)

    # За неделю
    today_week = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start_date_week = today_week - timedelta(days=today_week.weekday() + 4)
    end_date_week = start_date_week + timedelta(days=7)

    # За предыдущий месяц
    today_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_date_month = (today_month - timedelta(days=1)).replace(day=1)
    end_date_month = today_month

    # За месяц
    today_prev_month = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start_date_prev_month = today_prev_month.replace(day=1)
    end_date_prev_month = today_prev_month + timedelta(days=1)

    return {
        "start_date": start_date,
        "end_date": end_date,
        "start_date_week": start_date_week,
        "end_date_week": end_date_week,
        "start_date_month": start_date_month,
        "end_date_month": end_date_month,
        "start_date_prev_month": start_date_prev_month,
        "end_date_prev_month": end_date_prev_month
    }


@dp.callback_query_handler(text=["excel_users"], state=Analytics.main)
async def download_users_to_excel(call: types.CallbackQuery, state: FSMContext):
    """Выгрузка пользователей в Excel файл"""
    all_users_info = await db.get_all_users()
    result = []

    for user in all_users_info:
        created_at = user["created_at"].date()
        user_id = user["user_id"]
        username = user["username"]
        full_name = user["full_name"]
        gender = user["gender"]
        age_group = user["age_group"]
        referral = user["referral"]
        referral_id = user["referral_id"]
        card_fio = user["card_fio"]
        card_phone = user["card_phone"]
        card_number = user["card_number"]
        card_status = user["card_status"]
        birthday = user["birthday"]
        prize = user["prize"]
        balance = user["balance"]
        administrator = "Администратор" if user["administrator"] == True else None
        director = "Директор" if user["director"] == True else None
        ban_status = "Забанен" if user["ban_status"] == True else "Чист"
        reason_for_ban = user["reason_for_ban"]

        result.append(
            [username, full_name, created_at, birthday, user_id,  gender, age_group, card_phone, card_number, card_fio,
             card_status, prize, balance, referral, referral_id,
             administrator, director, ban_status, reason_for_ban]
        )

    book = openpyxl.Workbook()
    book.remove(book.active)
    book.create_sheet("Пользователи")
    sheet = book['Пользователи']

    sheet.cell(row=1, column=1).value = "Username"
    sheet.cell(row=1, column=2).value = "Полное имя"
    sheet.cell(row=1, column=3).value = "Дата регистрации"
    sheet.cell(row=1, column=4).value = "День рождения"
    sheet.cell(row=1, column=5).value = "User_id"
    sheet.cell(row=1, column=6).value = "Пол"
    sheet.cell(row=1, column=7).value = "Возрастная категория"
    sheet.cell(row=1, column=8).value = "Номер телефона"
    sheet.cell(row=1, column=9).value = "Номер карты"
    sheet.cell(row=1, column=10).value = "Фамилия Имя на карте"
    sheet.cell(row=1, column=11).value = "Статус карты"
    sheet.cell(row=1, column=12).value = "Призовые коды"
    sheet.cell(row=1, column=13).value = "Использовано призовых кодов"
    sheet.cell(row=1, column=14).value = "Referral_id"
    sheet.cell(row=1, column=15).value = "Referral"
    sheet.cell(row=1, column=16).value = "Администратор"
    sheet.cell(row=1, column=17).value = "Директор"
    sheet.cell(row=1, column=18).value = "Статус бана"
    sheet.cell(row=1, column=19).value = "Причина бана"

    fill_color = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    cell_range = sheet['A1':'S1']
    for row in cell_range:
        for cell in row:
            cell.fill = fill_color

    cell_range2 = tuple(sheet[get_column_letter(col) + '1'] for col in range(1, 20))
    for col_idx, column in enumerate(cell_range2, 1):
        column_letter = get_column_letter(col_idx)
        cells = list(sheet[column_letter])
        max_length = max(len(str(cell.value)) for cell in cells)
        adjusted_width = (max_length + 2) * 1.05
        sheet.column_dimensions[column_letter].width = adjusted_width

    for i, item in enumerate(result, start=2):

        sheet.cell(row=i, column=1).value = item[0]
        sheet.cell(row=i, column=2).value = item[1]
        sheet.cell(row=i, column=3).value = item[2]
        sheet.cell(row=i, column=4).value = item[3]
        sheet.cell(row=i, column=5).value = item[4]
        sheet.cell(row=i, column=6).value = item[5]
        sheet.cell(row=i, column=7).value = item[6]
        sheet.cell(row=i, column=8).value = item[7]
        sheet.cell(row=i, column=9).value = item[8]
        sheet.cell(row=i, column=10).value = item[9]
        sheet.cell(row=i, column=11).value = item[10]
        sheet.cell(row=i, column=12).value = item[11]
        sheet.cell(row=i, column=13).value = item[12]
        sheet.cell(row=i, column=14).value = item[13]
        sheet.cell(row=i, column=15).value = item[14]
        sheet.cell(row=i, column=16).value = item[15]
        sheet.cell(row=i, column=17).value = item[16]
        sheet.cell(row=i, column=18).value = item[17]
        sheet.cell(row=i, column=19).value = item[18]

    if not os.path.exists('temp'):
        os.mkdir('temp')
    book.save("temp/users.xlsx")
    book.close()

    with open("temp/users.xlsx", 'rb') as file:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Отправляем файл в сообщении
        xls_file = types.InputFile(output, filename="users.xlsx")

        msg = await bot.send_document(chat_id=call.message.chat.id, document=xls_file, caption="Выгрузка-Пользователи")

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)
    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.message_handler(Text(contains=["Пользователи"]), state=Analytics.main)
async def analytics_users(message: types.Message, state: FSMContext):
    """Ловлю нажатие на кнопку Пользователи"""
    await message.delete()
    users_nb_info = await db.get_all_nb_users()
    all_users_info = await db.get_all_users()

    text = f"Активных пользователей: {len(users_nb_info)}\nВсего пользователей: {len(all_users_info)}"
    markup = InlineKeyboardMarkup()
    markup.add(
        InlineKeyboardButton(text="Выгрузить в Excel файл", callback_data="excel_users")
    )
    msg = await message.answer(text=text, reply_markup=markup)

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)
    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.message_handler(Text(contains=["Рассылки"]), state=Analytics.main)
async def analytics_mailings(message: types.Message, state: FSMContext):
    """Ловлю нажатие на кнопку Формы рассылок-анкет"""
    await message.delete()
    total = 0
    type_mailing_list = ["standard", "birthday", "hall_reservation", "shipping", "loyal_card"]
    type_mailing_dict = {
        "standard": "📨 Обычная рассылка",
        "birthday": "🎁 Предложение для именинников",
        "hall_reservation": "🍽 Призыв к бронированию",
        "shipping": "🚚 Закажите доставку",
        "loyal_card": "💳 Владельцам карт лояльности"}
    text = ""
    kwargs = data_preparation()
    for item in type_mailing_list:
        # За сегодня
        today_count = len(
            await db.get_tasks_mailing(type_mailing=item, start_date=kwargs['start_date'], end_date=kwargs['end_date']))
        # За неделю
        week_count = len(
            await db.get_tasks_mailing(type_mailing=item, start_date=kwargs['start_date_week'],
                                       end_date=kwargs['end_date_week']))
        # За предыдущий месяц
        month_count = len(
            await db.get_tasks_mailing(type_mailing=item, start_date=kwargs['start_date_month'],
                                       end_date=kwargs['end_date_month']))
        # За месяц
        prev_month_count = len(
            await db.get_tasks_mailing(type_mailing=item,
                                       start_date=kwargs['start_date_prev_month'].replace(day=1) - timedelta(days=1),
                                       end_date=kwargs['end_date_prev_month']))

        text += f"<b>{type_mailing_dict[item]}</b>\n"
        text += f" За сегодня: {today_count}\n"
        text += f" За неделю: {week_count}\n"
        text += f" За текущий месяц: {prev_month_count}\n"
        text += f" За предыдущий месяц: {month_count}\n"
        text += f'{"-" * 50}\n'

    markup = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text='Показать активные рассылки', callback_data="show_active_tasks")]
        ]
    )

    msg = await message.answer(text=text, reply_markup=markup)

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)
    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.callback_query_handler(text_contains=["show_active_tasks"], state=Analytics.main)
async def get_all_active_tasks(call: types.CallbackQuery, state: FSMContext):
    """Обработка нажатия на кнопку показать активные рассылки"""
    data = call.data.split("-")
    if data[0][-3:] == 'off':
        task_id = data[-1]
        await db.off_task(task_id=task_id)

    all_active_tasks = await db.get_all_active_tasks()

    markup = InlineKeyboardMarkup()

    for task in all_active_tasks:
        markup.row(
            InlineKeyboardButton(text=f"{type_mailing_dict[task['type_mailing']]}",
                                 callback_data=f"view_task-{task['id']}"),
            InlineKeyboardButton(text=f"🚫 Отключить", callback_data=f"show_active_tasks_off-{task['id']}"),
        )
    if not all_active_tasks:
        text = "Активных рассылок нет"
    else:
        text = "Активные рассылки"

    if data[0][-3:] == 'off':
        with open('media/telegram.jpg', "rb") as file:
            photo = types.InputFile(file)

            await bot.edit_message_media(chat_id=call.from_user.id, message_id=call.message.message_id,
                                         media=InputMediaPhoto(photo), reply_markup=markup)

        await bot.edit_message_caption(chat_id=call.from_user.id, message_id=call.message.message_id, caption=text,
                                       reply_markup=markup)
    else:
        with open('media/telegram.jpg', "rb") as file:
            photo = types.InputFile(file)

            msg = await bot.send_photo(chat_id=call.from_user.id,
                                 photo=photo,
                                 caption=text, reply_markup=markup)

        data = await state.get_data()
        id_msg_list = data['id_msg_list']
        id_msg_list.append(msg.message_id)
        async with state.proxy() as data:
            data['id_msg_list'] = id_msg_list


@dp.callback_query_handler(text_contains=["view_task"], state=Analytics.main)
async def view_task_info(call: types.CallbackQuery, state: FSMContext):
    """Обработка нажатия на кнопку показать информацию о задании"""
    data = call.data.split("-")
    markup = call.message.reply_markup
    task_info = await db.get_task_info(task_id=data[-1])
    text = "Информация по подписке\n\n"
    text += f"Запланирована на: {datetime.strftime(task_info[0]['execution_date'], '%Y-%m-%d %H:%M')}\n"
    text += f"Тип рассылки: {type_mailing_dict[task_info[0]['type_mailing']]}\n"
    keyboard = {
        'hall_reservation': '🍽 Забронировать стол',
        'shipping': '🚚 Заказать доставку',
        'None': 'Никакой'
    }

    text += f"Кнопка: {keyboard[task_info[0]['keyboard']]}\n"
    text += f"{'-' * 50}\n"
    text += f"Текст: \n{task_info[0]['message']}\n"
    text += f"{'-' * 50}\n"
    text += f"Аминистратор: @{task_info[0]['admin_name']}\n\n"
    picture = task_info[0]['picture']
    await bot.edit_message_media(chat_id=call.from_user.id, message_id=call.message.message_id,
                                 media=InputMediaPhoto(picture))
    await bot.edit_message_caption(chat_id=call.from_user.id, message_id=call.message.message_id, caption=text,
                                   reply_markup=markup)


@dp.message_handler(Text(contains=["Статистика вызовова персонала"]), state=Analytics.main)
async def analytics_personal(message: types.Message, state: FSMContext):
    """Ловлю нажатие на кнопку Статистика вызовов персоналов"""
    await message.delete()
    # За сегодня
    kwargs = data_preparation()
    start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    who1 = "Официанта"
    who2 = "Кальянного мастера"
    today_count1 = len(await db.get_personal_request_today(personal=who1, start_date=kwargs['start_date'],
                                                           end_date=kwargs['end_date']))
    today_count2 = len(await db.get_personal_request_today(personal=who2, start_date=kwargs['start_date'],
                                                           end_date=kwargs['end_date']))

    text = "<b>За сегодня вызывали персонал</b>\n"
    text += f"{who1}: {plural_form(today_count1, 'раз')}\n"
    text += f"{who2}: {plural_form(today_count2, 'раз')}"

    msg = await message.answer(text=text)

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)
    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.message_handler(Text(contains=["Статистика бронирований"]), state=Analytics.main)
async def analytics_hall_reservation(message: types.Message, state: FSMContext):
    """Ловлю нажатие на кнопку Статистика бронирований"""
    await message.delete()
    # За сегодня
    kwargs = data_preparation()
    today_count = len(await db.get_approved_orders_hall(start_date=kwargs['start_date'], end_date=kwargs['end_date']))

    # За неделю
    week_count = len(
        await db.get_approved_orders_hall(start_date=kwargs['start_date_week'], end_date=kwargs['end_date_week']))

    # За предыдущий месяц
    month_count = len(
        await db.get_approved_orders_hall(start_date=kwargs['start_date_month'], end_date=kwargs['end_date_month']))

    # За месяц
    prev_month_count = len(
        await db.get_approved_orders_hall(start_date=kwargs['start_date_prev_month'].replace(day=1) - timedelta(days=1),
                                          end_date=kwargs['end_date_prev_month']))

    total = len(await db.get_all_approved_orders_hall())

    text = "<b>Бронирования (подтвержденные)</b>\n"
    text += f"За сегодня: {today_count}\n"
    text += f"За неделю: {week_count}\n"
    text += f"За текущий месяц: {prev_month_count}\n"
    text += f"За предыдущий месяц: {month_count}\n"
    text += f'{"-" * 50}\n'
    text += f"Всего: {total}\n"

    markup = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Показать бронирования сделанные сегодня (все)", callback_data="a_order_hall")],
            [InlineKeyboardButton(text="Показать бронирования на дату", callback_data="a_order_hall_data")]
        ]
    )

    msg = await message.answer(text=text, reply_markup=markup)

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)
    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.callback_query_handler(text=["a_order_hall_data"], state=Analytics.main)
async def get_order_hall_on_data(call: types.CallbackQuery, state: FSMContext):
    """Нажатие на кнопку Показать бронирования на дату"""
    data = await state.get_data()
    id_msg_list = data['id_msg_list']

    await Analytics.hall_reservation_statistic_data.set()

    date = datetime.now().strftime('%d.%m.%Y')
    text = f"Введите дату на которую нужно сделать выборку, в формате ДД.ММ.ГГГГ, сегодня {date}"

    msg = await call.message.answer(text=text)
    id_msg_list.append(msg.message_id)

    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.message_handler(content_types=["text"], state=Analytics.hall_reservation_statistic_data)
async def get_orders_on_date(message: types.Message, state: FSMContext):
    """Ловлю от пользователя дату выборки"""

    try:
        date = message.text.strip()
        if len(date.split('.')) == 3:
            if (len(date.split('.')[0]) == 2) and (len(date.split('.')[1]) == 2) and (len(date.split('.')[2]) == 4):
                orders = await db.get_orders_hall_on_date(date=datetime.strptime(date, "%d.%m.%Y"))

                data = await state.get_data()
                id_msg_list = data['id_msg_list']
                count = 0
                if orders:
                    for order in orders:
                        text = f"Бронирование столика на {order['data_reservation']} - {order['time_reservation']}\n"
                        text += f"Номер столика: {order['table_number']}\n"
                        text += f"Забронировал @{order['username']}\n"
                        text += f"Кол-во человек: {order['number_person']}\n"
                        text += f"Комментарий клиента: {order['comment']}\n"
                        text += f"Номер телефона: {order['phone']}\n"

                        if order['admin_answer'] == "approve":
                            markup = InlineKeyboardMarkup(
                                inline_keyboard=[
                                    [InlineKeyboardButton(text="Отменить",
                                                          callback_data=f"aa_hall_cancel-{order['id']}")]
                                ]
                            )
                        elif order['admin_answer'] == None or order['admin_answer'] == 'cancel' or order[
                            'admin_answer'] == 'rejected':
                            markup = InlineKeyboardMarkup(
                                inline_keyboard=[
                                    [
                                        InlineKeyboardButton(text="Взять в работу",
                                                             callback_data=f"aa_hall_approve-{order['id']}")
                                    ]
                                ]
                            )

                        msg = await message.answer(text=text, reply_markup=markup)
                        count += 1
                        id_msg_list.append(msg.message_id)
                else:
                    text = "На указанную Вами дату нет бронирований"
                    await message.answer(text=text)

                async with state.proxy() as data:
                    data['id_msg_list'] = id_msg_list

                await Analytics.main.set()
            else:
                raise Exception('input error')

        else:
            raise Exception("input error")
    except Exception as _ex:
        text = ""
        if (str(_ex) == 'input error') or (str(_ex) == 'day is out of range for month'):
            text = f"К сожалению я Вас не понимаю, введите корректную дату в правильном формате ДД.ММ.ГГГГ. Сегодня {datetime.strftime(datetime.now(), '%d.%m.%Y')}"

        await message.answer(text=text)
        return


@dp.callback_query_handler(text_contains=["a_order_hall"], state=Analytics.main)
async def get_order_hall_made_today(call: types.CallbackQuery, state: FSMContext):
    """Выборка бронирований столиков сделанных сегодня"""
    orders = await db.get_all_order_hall_made_today(date=datetime.now().date())
    data = await state.get_data()
    id_msg_list = data['id_msg_list']

    for order in orders:
        text = f"Бронирование столика на {order['data_reservation']} - {order['time_reservation']}\n"
        text += f"Номер столика: {order['table_number']}\n"
        text += f"Забронировал @{order['username']}\n"
        text += f"Кол-во человек: {order['number_person']}\n"
        text += f"Комментарий клиента: {order['comment']}\n"
        text += f"Номер телефона: {order['phone']}\n"

        if order['admin_answer'] == "approve":
            markup = InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text="Отменить", callback_data=f"aa_hall_cancel-{order['id']}")]
                ]
            )
        elif order['admin_answer'] == None or order['admin_answer'] == 'cancel':
            markup = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(text="Взять в работу", callback_data=f"aa_hall_approve-{order['id']}")
                    ]
                ]
            )
        msg = await call.message.answer(text=text, reply_markup=markup)
        id_msg_list.append(msg.message_id)

    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.callback_query_handler(text_contains=["aa_hall_cancel"], state=Analytics.main)
async def cancel_order_hall_made_today(call: types.CallbackQuery):
    """Отмена бронирования администратором"""

    data = call.data.split('-')

    markup = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Взять в работу",
                                  callback_data=f"aa_hall_approve-{data[1]}")]
        ]
    )

    await db.update_order_hall_status(id=int(data[1]), order_status=True, admin_answer=data[0].split("_")[-1],
                                      admin_id=str(call.from_user.id), admin_name=call.from_user.username,
                                      table_number=0)
    order = await db.get_order_hall_data(id=int(data[1]))

    text = f"Бронирование столика на {order[0]['data_reservation']} - {order[0]['time_reservation']}\n"
    text += f"Номер столика: {order[0]['table_number']}\n"
    text += f"Забронировал @{order[0]['username']}\n"
    text += f"Кол-во человек: {order[0]['number_person']}\n"
    text += f"Комментарий клиента: {order[0]['comment']}\n"
    text += f"Номер телефона: {order[0]['phone']}\n"

    await call.message.edit_text(text=text)
    await call.message.edit_reply_markup(markup)


@dp.callback_query_handler(text_contains=["aa_hall_approve"], state=Analytics.main)
async def approve_order_hall_made_today(call: types.CallbackQuery, state: FSMContext):
    """Взять заявку на бронирование в работу администратором"""
    data = call.data.split('-')
    order_id = int(data[1])
    markup = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Отменить", callback_data=f"aa_hall_cancel-{data[1]}")]
        ]
    )
    await call.message.edit_reply_markup(markup)

    text = "Выедите номер столика"
    msg = await call.message.answer(text=text)

    await Analytics.hall_reservation_statistic_table.set()
    async with state.proxy() as data:
        data['order_id'] = order_id
        data['message_id'] = call.message.message_id
        data['markup'] = markup


@dp.message_handler(content_types=["text"], state=Analytics.hall_reservation_statistic_table)
async def get_orders_on_date(message: types.Message, state: FSMContext):
    """Лобвлю номер столика от администратора"""
    data = await state.get_data()

    await db.update_order_hall_status(id=int(data['order_id']), order_status=False, admin_answer='approve',
                                      admin_id=str(message.from_user.id), admin_name=message.from_user.username,
                                      table_number=int(message.text))

    order = await db.get_order_hall_data(id=int(data['order_id']))

    text = f"Бронирование столика на {order[0]['data_reservation']} - {order[0]['time_reservation']}\n"
    text += f"Номер столика: {order[0]['table_number']}\n"
    text += f"Забронировал @{order[0]['username']}\n"
    text += f"Кол-во человек: {order[0]['number_person']}\n"
    text += f"Комментарий клиента: {order[0]['comment']}\n"
    text += f"Номер телефона: {order[0]['phone']}\n"

    await bot.edit_message_text(chat_id=message.from_user.id, message_id=int(data['message_id']), text=text)
    await bot.edit_message_reply_markup(chat_id=message.from_user.id, message_id=int(data['message_id']),
                                        reply_markup=data['markup'])

    await bot.send_message(chat_id=message.from_user.id, text="Заявка успешно изменена",
                           reply_to_message_id=int(data['message_id']))
    await Analytics.main.set()


@dp.message_handler(Text(contains=["Статистика доставки"]), state=Analytics.main)
async def analytics_shipping(message: types.Message, state: FSMContext):
    """Ловлю нажатие на кнопку Статистика доставки"""
    await message.delete()
    kwargs = data_preparation()
    # За сегодня
    today_count = len(await db.get_approved_shipping(start_date=kwargs['start_date'], end_date=kwargs['end_date']))

    # За неделю
    week_count = len(
        await db.get_approved_shipping(start_date=kwargs['start_date_week'], end_date=kwargs['end_date_week']))

    # За предыдущий месяц
    month_count = len(
        await db.get_approved_shipping(start_date=kwargs['start_date_month'], end_date=kwargs['end_date_month']))

    # За месяц
    prev_month_count = len(
        await db.get_approved_shipping(start_date=kwargs['start_date_prev_month'].replace(day=1) - timedelta(days=1),
                                       end_date=kwargs['end_date_prev_month']))

    total = len(await db.get_all_approved_shipping())

    text = "<b>Заказы на доставку (подтвержденные)</b>\n"
    text += f"За сегодня: {today_count}\n"
    text += f"За неделю: {week_count}\n"
    text += f"За текущий месяц: {prev_month_count}\n"
    text += f"За предыдущий месяц: {month_count}\n"
    text += f'{"-" * 50}\n'
    text += f"Всего: {total}\n"

    markup = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Выгрузить в Excel", callback_data="a_sh_excel")],
            [InlineKeyboardButton(text="Показать заявки сделанные сегодня (все)", callback_data="a_sh_order")]
        ]
    )

    msg = await message.answer(text=text, reply_markup=markup)

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)
    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


def paste_data_to_table(sheet, order):
    """Вставка данных в таблицу"""
    for i, order in enumerate(order, start=2):
        sheet.cell(row=i, column=1).value = order["user_name"]
        sheet.cell(row=i, column=2).value = order["user_id"]
        sheet.cell(row=i, column=3).value = order["phone"]
        sheet.cell(row=i, column=4).value = order["data_reservation"]
        sheet.cell(row=i, column=5).value = order["time_reservation"]
        sheet.cell(row=i, column=6).value = order["address"]
        sheet.cell(row=i, column=7).value = "Карта" if order["pay_method"] == "pay_method_card" else "Наличные"
        sheet.cell(row=i, column=8).value = order["number_of_devices"]
        zakaz = json.loads(order["tpc"])
        order_text = ""
        for item in zakaz:
            order_text += f"{item['title']} - кол-во порций {item['count']}\n"
        sheet.cell(row=i, column=9).value = order_text
        sheet.cell(row=i, column=10).value = order["final_summa"]


@dp.callback_query_handler(text=["a_sh_excel"], state=Analytics.main)
async def export_shipping_to_excel(call: types.CallbackQuery, state: FSMContext):
    """Выгрузка данных по доставке в Excell"""

    kwargs = data_preparation()
    # За сегодня
    today = await db.get_approved_shipping(start_date=kwargs['start_date'], end_date=kwargs['end_date'])
    # За неделю
    week = await db.get_approved_shipping(start_date=kwargs['start_date_week'], end_date=kwargs['end_date_week'])
    # За месяц
    month = await db.get_approved_shipping(
        start_date=kwargs['start_date_prev_month'].replace(day=1) - timedelta(days=1),
        end_date=kwargs['end_date_prev_month'])
    # За предыдущий месяц
    prev_month = await db.get_approved_shipping(start_date=kwargs['start_date_month'],
                                                end_date=kwargs['end_date_month'])

    book = openpyxl.Workbook()
    book.remove(book.active)
    book.create_sheet("За сегодня")
    book.create_sheet("За неделю")
    book.create_sheet("За месяц")
    book.create_sheet("За предыдущий месяц")

    for sheet in book.worksheets:
        sheet.cell(row=1, column=1).value = 'Username'
        sheet.cell(row=1, column=2).value = 'User_id'
        sheet.cell(row=1, column=3).value = 'Телефон'
        sheet.cell(row=1, column=4).value = 'Дата'
        sheet.cell(row=1, column=5).value = 'Время'
        sheet.cell(row=1, column=6).value = 'Адрес доставки'
        sheet.cell(row=1, column=7).value = 'Метод оплаты'
        sheet.cell(row=1, column=8).value = 'Кол-во приборов'
        sheet.cell(row=1, column=9).value = 'Заказ'
        sheet.cell(row=1, column=10).value = 'Общая сумма'

        fill_color = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
        cell_range = sheet['A1':'J1']
        for row in cell_range:
            for cell in row:
                cell.fill = fill_color

        cell_range2 = tuple(sheet[get_column_letter(col) + '1'] for col in range(1, 11))
        for col_idx, column in enumerate(cell_range2, 1):
            column_letter = get_column_letter(col_idx)
            cells = list(sheet[column_letter])
            max_length = max(len(str(cell.value)) for cell in cells)
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        if sheet.title == "За сегодня":
            paste_data_to_table(sheet=sheet, order=today)
            if len(today) > 0:
                sheet.cell(row=len(today) + 3, column=10).value = f"=SUM(J2:J{len(today) + 1})"
                sheet.cell(row=len(today) + 3, column=9).value = "ИТОГО:"
                fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                sheet.cell(row=len(today) + 3, column=9).fill = fill
                sheet.cell(row=len(today) + 3, column=10).fill = fill
        elif sheet.title == "За неделю":
            paste_data_to_table(sheet=sheet, order=week)
            if len(week) > 0:
                sheet.cell(row=len(week) + 3, column=10).value = f"=SUM(J2:J{len(week) + 1})"
                sheet.cell(row=len(week) + 3, column=9).value = "ИТОГО:"
                fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                sheet.cell(row=len(week) + 3, column=9).fill = fill
                sheet.cell(row=len(week) + 3, column=10).fill = fill
        elif sheet.title == "За месяц":
            paste_data_to_table(sheet=sheet, order=month)
            if len(month) > 0:
                sheet.cell(row=len(month) + 3, column=10).value = f"=SUM(J2:J{len(month) + 1})"
                sheet.cell(row=len(month) + 3, column=9).value = "ИТОГО:"
                fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                sheet.cell(row=len(month) + 3, column=9).fill = fill
                sheet.cell(row=len(month) + 3, column=10).fill = fill
        elif sheet.title == "За предыдущий месяц":
            paste_data_to_table(sheet=sheet, order=prev_month)
            if len(prev_month) > 0:
                sheet.cell(row=len(prev_month) + 3, column=10).value = f"=SUM(J2:J{len(prev_month) + 1})"
                sheet.cell(row=len(prev_month) + 3, column=9).value = "ИТОГО:"
                fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                sheet.cell(row=len(prev_month) + 3, column=9).fill = fill
                sheet.cell(row=len(prev_month) + 3, column=10).fill = fill
                sheet.cell(row=len(prev_month) + 3, column=10).fill = fill


    if not os.path.exists('temp'):
        os.mkdir('temp')

    book.save("temp/shipping.xlsx")
    book.close()

    with open("temp/shipping.xlsx", 'rb') as file:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Отправляем файл в сообщении
        xls_file = types.InputFile(output, filename="shipping.xlsx")

        msg = await bot.send_document(chat_id=call.message.chat.id, document=xls_file, caption="Выгрузка-Доставка")

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)
    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.callback_query_handler(text=["a_sh_order"], state=Analytics.main)
async def get_shipping_order_made_today(call: types.CallbackQuery, state: FSMContext):
    """Выборка заявок на доставку сделанных сегодня"""
    orders = await db.get_shipping_order_made_today(date=datetime.now().date())
    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    if orders:
        for item in orders:
            text = f"Заявка на {item['data_reservation']} {item['time_reservation']}\n"
            text += f"от @{item['user_name']}\n"
            text += f"Адрес: {item['address']}\n"
            text += f"Телофон: {item['phone']}\n"
            text += f"Кол-во приборов: {item['number_of_devices']}\n"
            text += f"{'-' * 50}\n"
            data = json.loads(item['tpc'])
            for title in data:
                text += f"{title['title']} - {title['count']}\n"
            text += f'{"-" * 50}\n'
            text += f"Итого: {item['final_summa']} тенге"
            if item['admin_answer'] == "approve":
                markup = InlineKeyboardMarkup(
                    inline_keyboard=[
                        [InlineKeyboardButton(text="Отменить", callback_data=f"aa_sh_cancel-{item['id']}")]
                    ]
                )
            elif item['admin_answer'] == 'cancel':
                markup = InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(text="Взять в работу", callback_data=f"aa_sh_approve-{item['id']}")
                        ]
                    ]
                )
            msg = await call.message.answer(text=text, reply_markup=markup)
            id_msg_list.append(msg.message_id)
    else:
        await call.message.answer(text="Сегодня еще не было заказов на доставку")
    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list


@dp.callback_query_handler(text_contains=["aa_sh_"], state=Analytics.main)
async def approve_shipping_order_made_today(call: types.CallbackQuery, state: FSMContext):
    """Подтверждение/Отмена заявки администратором"""
    username = re.findall(r'@(\w+)', call.message.html_text)[0]
    user = await db.get_user_by_username(username=username)

    data = call.data.split('-')
    if 'approve' in data[0]:
        markup = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="Отменить", callback_data=f"aa_sh_cancel-{data[1]}")]
            ]
        )
        order_status = False
        await bot.send_message(chat_id=int(user[0]['user_id']),
                               text=f'Ваша заявка на доставку подтверждена администрацией.\n\n{call.message.text}')

    elif 'cancel' in data[0]:
        markup = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(text="Взять в работу", callback_data=f"aa_sh_approve-{data[1]}")
                ]
            ]
        )
        order_status = True
        await bot.send_message(chat_id=int(user[0]['user_id']),
                               text=f'Ваша заявка на доставку отклонена администрацией.\n\n{call.message.text}')

    await call.message.edit_reply_markup(markup)
    await db.update_shipping_order_status(id=int(data[1]), admin_name=call.from_user.username,
                                          admin_id=str(call.from_user.id), admin_answer=data[0].split("_")[-1],
                                          order_status=order_status)


@dp.message_handler(Text(contains=["Участники программы лояльности"]), state=Analytics.main)
async def analytics_loyal_program_participants(message: types.Message, state: FSMContext):
    """Ловлю нажатие на кнопку Участники программы лояльности"""
    await message.delete()
    loyal_program_participants_info = await db.get_loyal_program_participants()
    markup = InlineKeyboardMarkup()
    for user_info in loyal_program_participants_info:
        markup.add(
            InlineKeyboardButton(text=user_info["username"],
                                 callback_data=f"{user_info['user_id']}-user_info")
        )
    msg = await message.answer(text="Участники программы лояльности", reply_markup=markup)

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)

    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list
        data["markup"] = markup


@dp.callback_query_handler(text_contains=["user_info"], state=Analytics.main)
async def show_user_info(call: types.CallbackQuery, state: FSMContext):
    """Показать информацию о пользователе"""
    # await bot.answer_callback_query(call.id, text="Привет, это alert!", show_alert=True)
    callback_data = call.data.split('-')
    data = await state.get_data()
    user_info = await db.get_user_info(user_id=callback_data[0])
    all_invited_users = await db.get_all_invited_users(user_info[0]["referral_id"])

    count_users = len(all_invited_users)
    created_at = user_info[0]['created_at']

    text = f"Информаци о пользователе @{user_info[0]['username']}\n\n"
    text += f"Дата регистрации: {datetime.strftime(created_at, '%Y-%m-%d')}\n"
    text += f"Фамилия и Имя: {user_info[0]['card_fio']}\n"
    text += f"Номер карты: {user_info[0]['card_number']}\n"
    text += f"Дата рождения: {user_info[0]['birthday']}\n"
    text += f"Пришло к нам по рекомендации: {plural_form(count_users, 'человек')}\n"
    text += f"Номер телефона: {user_info[0]['card_phone']}\n"
    text += f"{'-' * 50}\n"
    text += f"Бан Статус: {'<b>Забанен</b>' if user_info[0]['ban_status'] == True else 'Чист'}\n"
    if user_info[0]["ban_status"] == True:
        text += f"Причина бана: {user_info[0]['reason_for_ban']}\n"
        text += f"Дата получения бана: {user_info[0]['ban_date']}"

    msg = await call.message.edit_text(text=text)
    await call.message.edit_reply_markup(data["markup"])

    data = await state.get_data()
    id_msg_list = data['id_msg_list']
    id_msg_list.append(msg.message_id)

    async with state.proxy() as data:
        data['id_msg_list'] = id_msg_list
